import smtplib
import random
from datetime import *
import cv2
import numpy as np
import os
from openpyxl import load_workbook
from pytz import timezone
import tensorflow as tf


relay = 21

# import RPi.GPIO as GPIO
# import time
# GPIO.setmode(GPIO.BCM)
# GPIO.setwarnings(False)
# GPIO.setup(relay, GPIO.OUT)

wb = load_workbook('personal_data.xlsx')
ws = wb.active

my_model = tf.keras.models.load_model('model.h5')
emotion_dict = {0: "Angry", 1: "Disgusted", 2: "Fearful", 3: "Happy", 4: "Neutral", 5: "Sad", 6: "Surprised"}

haar = 'haarcascade_frontalface_default.xml'
data = ['level_1', 'level_2']

print('Recognizing face please be in sufficient lights...')

(images, lables, names, id) = ([], [], {}, 0)

for d in data:
    for (subdirs, dirs, files) in os.walk(d):
        for subdir in dirs:
            names[id] = subdir
            subjectpath = os.path.join(d, subdir)
            for filename in os.listdir(subjectpath):
                path = subjectpath + '/' + filename
                lable = id
                images.append(cv2.imread(path, 0))
                lables.append(int(lable))
            id += 1
(width, height) = (130, 100)
(images, lables) = [np.array(lis) for lis in [images, lables]]

model = cv2.face.LBPHFaceRecognizer_create()
model.train(images, lables)

face_cascade = cv2.CascadeClassifier(haar)
webcam = cv2.VideoCapture(0)    # RPi video feed

while True:
    (_, im) = webcam.read()
    gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break
    faces = face_cascade.detectMultiScale(gray, 1.3, 5)
    for (x, y, w, h) in faces:
        cv2.rectangle(im, (x, y), (x + w, y + h), (255, 0, 0), 2)
        face = gray[y:y + h, x:x + w]
        face_resize = cv2.resize(face, (width, height))
        pred = model.predict(face_resize)
        cv2.rectangle(im, (x, y), (x + w, y + h), (0, 255, 0), 3)

        if pred[1] < 100:
            cropped_img = np.expand_dims(np.expand_dims(cv2.resize(face_resize, (48, 48)), -1), 0)
            prediction = my_model.predict(cropped_img)
            maxindex = int(np.argmax(prediction))
            print(emotion_dict[maxindex])
            if emotion_dict[maxindex] in ['Happy', 'Neutral']:
                now = datetime.now(timezone('UTC'))
                now_asia = now.astimezone(timezone('Asia/Kolkata'))

                if names[pred[0]] in os.listdir('level_1'):
                    k = 0
                    a = 1
                    for cell in ws['A']:
                        if cell.value is not None:
                            if names[pred[0]] in cell.value:
                                r = cell.row

                    r1 = 'B' + str(r)
                    r2 = 'C' + str(r)

                    while k < 3:
                        p = input('Enter your password: ')

                        if p == ws[r1].value:
                            j = 0
                            otp = random.randint(1000, 9999)

                            server = smtplib.SMTP('smtp.gmail.com', port=587)
                            server.starttls()
                            server.login("gajeelredfox746@gmail.com", "passWORD123")

                            msg = str(otp)

                            server.sendmail("gajeelredfox746@gmail.com", ws[r2].value, msg)

                            server.quit()

                            print('Check your mail for OTP')

                            while j < 3:
                                otp1 = int(input('Enter the OTP: '))
                                if otp1 == otp:
                                    print('Level 1 access granted')
                                    print('Welcome ', names[pred[0]])

                                    # GPIO.output(relay, True)
                                    # time.sleep(3)
                                    # GPIO.output(relay, False)
                                    # time.sleep(3)

                                    wb1 = load_workbook('entry_log.xlsx')
                                    ws1 = wb1.active

                                    r2 = ws1.max_row + 1
                                    ws1.cell(column=1, row=r2, value=names[pred[0]])
                                    ws1.cell(column=2, row=r2, value=now_asia.strftime("%d/%m/%Y"))
                                    ws1.cell(column=3, row=r2, value=now_asia.strftime("%H:%M:%S"))

                                    wb1.save(filename='entry_log.xlsx')
                                    wb1.close()

                                else:
                                    if j != 2:
                                        print('Incorrect OTP')
                                        j += 1
                                    else:
                                        exit()
                            break

                        else:
                            if k != 2:
                                print('Incorrect password')
                                k += 1
                            else:
                                exit()

                else:
                    a = 2
                    k1 = 0
                    for cell in ws['A']:
                        if cell.value is not None:
                            if names[pred[0]] in cell.value:
                                r = cell.row

                    r1 = 'B' + str(r)
                    r2 = 'C' + str(r)

                    while k1 < 3:
                        p = input('Enter your password: ')

                        if p == ws[r1].value:
                            print('Level 2 access granted')
                            print('Welcome ', names[pred[0]])

                            # GPIO.output(relay, True)
                            # time.sleep(3)
                            # GPIO.output(relay, False)
                            # time.sleep(3)

                            wb1 = load_workbook('entry_log.xlsx')
                            ws1 = wb1.active

                            r2 = ws1.max_row + 1
                            ws1.cell(column=1, row=r2, value=names[pred[0]])
                            ws1.cell(column=2, row=r2, value=now_asia.strftime("%d/%m/%Y"))
                            ws1.cell(column=3, row=r2, value=now_asia.strftime("%H:%M:%S"))

                            wb1.save(filename='entry_log.xlsx')
                            wb1.close()


                        else:
                            print('Incorrect Password')
                            k1 += 1

                cv2.putText(im, '%s - level %s- %.0f' % (names[pred[0]], a, pred[1]), (x - 10, y - 10), cv2.FONT_HERSHEY_PLAIN, 1, (0, 255, 0))

            else:
                cv2.putText(im, emotion_dict[maxindex], (x - 10, y - 10), cv2.FONT_HERSHEY_PLAIN, 1, (0, 255, 0))
                print('Personnel seems ' + emotion_dict[maxindex] + '. Not safe to give access')

        else:
            cv2.putText(im, 'not recognized', (x - 10, y - 10), cv2.FONT_HERSHEY_PLAIN, 1, (0, 255, 0))
            print('Not recognized. Please try again.')

    cv2.imshow('OpenCV', im)

    key = cv2.waitKey(10)
    if key == 27:
        break

