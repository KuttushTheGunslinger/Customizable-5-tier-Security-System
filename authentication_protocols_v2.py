import cv2
import numpy as np
import os
import random
import smtplib
import tensorflow as tf
from datetime import *
from openpyxl import load_workbook
from pytz import timezone
import tkinter as tk
from tkinter import *
import time
# import hashlib
# from pyfingerprint.pyfingerprint import PyFingerprint
# from pyfingerprint.pyfingerprint import FINGERPRINT_CHARBUFFER1
pin = ''
tries, tries1 = 0, 0
pin1 = ''


def security():
    wb = load_workbook('personal_data.xlsx')
    ws = wb.active

    wb1 = load_workbook('system_info.xlsx')
    ws1 = wb1.active

    haar = 'haarcascade_frontalface_default.xml'
    data = ['level_1', 'level_2']
    (width, height) = (130, 100)

    def password(n):
        k = 0
        a = 1
        for cell in ws['A']:
            if cell.value is not None:
                if n in cell.value:
                    r = cell.row

        r1 = 'B' + str(r)
        password_input = ws[r1].value

        def code1(value1, password_input):
            global pin1
            global tries1
            if value1 == 'X':
                pin1 = pin1[:-1]
                e1.delete('0', 'end')
                e1.insert('end', pin1)

            elif value1 == 'Enter':
                if pin1 == password_input:
                    root1.destroy()

                else:
                    Label4 = tk.Label(root1, text='PIN ERROR')
                    Label4.place(relx=0.0, rely=1.0, anchor='sw')
                    tries1 = tries1 + 1
                    Label3 = tk.Label(root1, text='Attempts left: ' + str(3 - tries1))
                    Label3.place(relx=1.0, rely=0.0, anchor='ne')
                    if tries1 < 3:
                        pin1 = ''
                        e1.delete('0', 'end')
                    else:
                        root1.destroy()
                        exit()

            else:
                pin1 += value1
                e1.insert('end', value1)

        keys1 = [
            ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h'],
            ['i', 'j', 'k', 'l', 'n', 'o', 'p', 'q'],
            ['r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z'],
            ['1', '2', '3', '4', '5', '6', '7', '8', '9'],
            ['X', '0', 'Enter'],
        ]

        root1 = tk.Tk()

        root1.geometry("480x320")

        e1 = tk.Entry(root1)
        e1.grid(row=0, column=0, columnspan=3, ipady=5)

        for y12, row in enumerate(keys1, 1):
            for x12, key1 in enumerate(row):
                bpass = tk.Button(root1, text=key1, command=lambda val=key1: code1(val, password_input))
                bpass.grid(row=y12, column=x12, ipadx=10, ipady=10)

        root1.mainloop()

    def otp(n):
        for cell in ws['A']:
            if cell.value is not None:
                if n in cell.value:
                    r = cell.row

        r2 = 'C' + str(r)

        o = str(random.randint(1000, 9999))

        server = smtplib.SMTP('smtp.gmail.com', port=587)
        server.starttls()
        server.login("gajeelredfox746@gmail.com", "passWORD123")

        msg = o

        server.sendmail("gajeelredfox746@gmail.com", ws[r2].value, msg)

        server.quit()

        print('Check your mail for OTP')

        def code(value, o):
            global pin
            global tries
            if value == 'X':
                pin = pin[:-1]
                e.delete('0', 'end')
                e.insert('end', pin)

            elif value == 'Enter':
                if pin == o:
                    root.destroy()

                else:
                    Label2 = tk.Label(root, text='PIN ERROR')
                    Label2.place(relx=0.0, rely=1.0, anchor='sw')
                    tries = tries + 1
                    Label1 = tk.Label(root, text='Attempts left: ' + str(3 - tries))
                    Label1.place(relx=1.0, rely=0.0, anchor='ne')
                    if tries < 3:
                        pin = ''
                        e.delete('0', 'end')
                    else:
                        root.destroy()
                        exit()
            else:
                pin += value
                e.insert('end', value)

        keys = [
            ['1', '2', '3'],
            ['4', '5', '6'],
            ['7', '8', '9'],
            ['X', '0', 'Enter'],
        ]

        root = tk.Tk()

        root.geometry("480x320")

        e = tk.Entry(root)
        e.grid(row=0, column=0, columnspan=3, ipady=5)

        for y111, row in enumerate(keys, 1):
            for x111, key in enumerate(row):
                b = tk.Button(root, text=key, command=lambda val=key: code(val, o))
                b.grid(row=y111, column=x111, ipadx=10, ipady=10)

        root.mainloop()

    # def fingerprint_testing(Name):
    #     fp_count = 3
    #     l = []
    #     try:
    #         f = PyFingerprint('/dev/ttyUSB0', 57600, 0xFFFFFFFF, 0x00000000)
    #
    #         if ( f.verifyPassword() == False ):
    #             raise ValueError('The given fingerprint sensor password is wrong!')
    #
    #     except Exception as e:
    #         print('The fingerprint sensor could not be initialized!')
    #         print('Exception message: ' + str(e))
    #         exit(1)
    #
    #     ## Gets some sensor information
    #     print('Currently used templates: ' + str(f.getTemplateCount()) +'/'+ str(f.getStorageCapacity()))
    #
    #     ## Tries to search the finger and calculate hash
    #     try:
    #         while fp_count < 3:
    #             print('Waiting for finger...')
    #
    #             ## Wait that finger is read
    #             while ( f.readImage() == False ):
    #                 pass
    #
    #             ## Converts read image to characteristics and stores it in charbuffer 1
    #             f.convertImage(FINGERPRINT_CHARBUFFER1)
    #
    #             ## Searchs template
    #             result = f.searchTemplate()
    #
    #             positionNumber = result[0]
    #             accuracyScore = result[1]
    #
    #             if fp_count < 3:
    #                 if ( positionNumber == -1 ):
    #                     print('No match found!')
    #                     fp_count = fp_count + 1
    #
    #                 else:
    #                     wb3 = load_workbook('fingerprints.xlsx')
    #                     ws3 = wb3.active
    #
    #                     for cell in ws3['A']:
    #                         l.append(cell.value)
    #
    #                     if Name in l:
    #                         r = l.index(Name) + 1
    #                         print(r)
    #
    #                     if positionNumber == r - 2:
    #                         print("Fingerprint matched")
    #                         print('Found template at position #' + str(positionNumber))
    #                         print('The accuracy score is: ' + str(accuracyScore))
    #                         fp_count = 3
    #             else:
    #                 exit()
    #
    #         else:
    #             exit()
    #
    #         wb3.save(filename='fingerprints.xlsx')
    #         wb3.close()
    #
    #     except Exception as e:
    #         print('Operation failed!')
    #         print('Exception message: ' + str(e))
    #         exit(1)

    def emotion():
        cap = cv2.VideoCapture(0)
        ret, image = cap.read()
        c = 1
        while c:
            gray1 = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            faces1 = face_cascade.detectMultiScale(gray1, 1.3, 5)
            for (x1, y1, w1, h1) in faces1:
                cv2.rectangle(image, (x1, y1), (x1 + w1, y1 + h1), (255, 0, 0), 2)
                face1 = gray1[y1:y1 + h1, x1:x1 + w1]
                face_resize1 = cv2.resize(face1, (width, height))
                cv2.imshow("face_resize1", face_resize1)
                cropped_img = np.expand_dims(np.expand_dims(cv2.resize(face_resize1, (48, 48)), -1), 0)

                my_model = tf.keras.models.load_model('model.h5')
                emotion_dict = {0: "Angry", 1: "Disgusted", 2: "Fearful", 3: "Happy", 4: "Neutral", 5: "Sad", 6: "Surprised"}
                prediction = my_model.predict(cropped_img)
                maxindex = int(np.argmax(prediction))
                print('hi')
                if emotion_dict[maxindex] in ['Happy', 'Neutral']:
                    print("Personnel is ", emotion_dict[maxindex])
                    c = 0
                    cap.release()
                else:
                    cv2.putText(im, emotion_dict[maxindex], (x - 10, y - 10), cv2.FONT_HERSHEY_PLAIN, 1, (0, 255, 0))
                    print('Personnel seems ' + emotion_dict[maxindex] + '. Not safe to give access')

    (images, lables, names, id) = ([], [], {}, 0)

    for d in data:
        for (subdirs, dirs, files) in os.walk(d):
            for subdir in dirs:
                names[id] = subdir
                subjectpath = os.path.join(d, subdir)
                for filename in os.listdir(subjectpath):
                    path = subjectpath + '/' + filename
                    lable = id
                    images.append(cv2.imread(path, 0))
                    lables.append(int(lable))
                id += 1
    (images, lables) = [np.array(lis) for lis in [images, lables]]

    model = cv2.face.LBPHFaceRecognizer_create()
    model.train(images, lables)

    face_cascade = cv2.CascadeClassifier(haar)
    top.destroy()

    while True:
        webcam = cv2.VideoCapture(0)
        (_, im) = webcam.read()
        gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)
        for (x, y, w, h) in faces:
            cv2.rectangle(im, (x, y), (x + w, y + h), (255, 0, 0), 2)
            face = gray[y:y + h, x:x + w]
            face_resize = cv2.resize(face, (width, height))
            cv2.imshow("frame1", im)
            pred = model.predict(face_resize)
            print(pred)
            cv2.rectangle(im, (x, y), (x + w, y + h), (0, 255, 0), 3)
            webcam.release()
            if pred[1] < 100:
                print('Face recognized.')
                print('Hello ', names[pred[0]])
                if names[pred[0]] in os.listdir('level_1'):
                    if ws1['B2'].value == 1:
                        emotion()
                    # if ws1['B3'].value == 1:
                    #     fingerprint_testing(names[pred[0]])
                    if ws1['B4'].value == 1:
                        password(names[pred[0]])
                    if ws1['B5'].value == 1:
                        otp(names[pred[0]])

                    now = datetime.now(timezone('UTC'))
                    now_asia = now.astimezone(timezone('Asia/Kolkata'))
                    wb2 = load_workbook('entry_log.xlsx')
                    ws2 = wb2.active

                    r3 = ws2.max_row + 1
                    ws2.cell(column=1, row=r3, value=names[pred[0]])
                    ws2.cell(column=2, row=r3, value=now_asia.strftime("%d/%m/%Y"))
                    ws2.cell(column=3, row=r3, value=now_asia.strftime("%H:%M:%S"))

                    wb2.save(filename='entry_log.xlsx')
                    wb2.close()
                    exit()

                else:
                    if ws1['B8'].value == 1:
                        emotion()
                    # if ws1['B9'].value == 1:
                    #     fingerprint_testing(names[pred[0]])
                    if ws1['B10'].value == 1:
                        password(names[pred[0]])
                    if ws1['B11'].value == 1:
                        otp(names[pred[0]])

                    now = datetime.now(timezone('UTC'))
                    now_asia = now.astimezone(timezone('Asia/Kolkata'))
                    wb2 = load_workbook('entry_log.xlsx')
                    ws2 = wb2.active

                    r3 = ws2.max_row + 1
                    ws2.cell(column=1, row=r3, value=names[pred[0]])
                    ws2.cell(column=2, row=r3, value=now_asia.strftime("%d/%m/%Y"))
                    ws2.cell(column=3, row=r3, value=now_asia.strftime("%H:%M:%S"))

                    wb2.save(filename='entry_log.xlsx')
                    wb2.close()
                    exit()


top = tk.Tk()

top.geometry("480x320")

B = tk.Button(top, text="Press to enter", command=security)
B.pack()
top.mainloop()
