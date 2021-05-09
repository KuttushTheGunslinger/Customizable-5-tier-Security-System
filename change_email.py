import smtplib
import random
from openpyxl import load_workbook

i, j = 0, 0

wb = load_workbook('personal_data.xlsx')
ws = wb.active

n = input('Enter your name: ')
l = []
for cell in ws['A']:
    l.append(cell.value)

if n in l:
    r = l.index(n) + 1
else:
    print("User doesn't exist")
    exit()

r1 = 'C' + str(r)

if ws[r1].value is not None:
    while i < 3:
        p = input('Enter your current email: ')
        if p == ws[r1].value:
            o = random.randint(1000, 9999)
            server = smtplib.SMTP('smtp.gmail.com', port=587)
            server.starttls()
            server.login("gajeelredfox746@gmail.com", "passWORD123")

            msg = str(o)

            server.sendmail("gajeelredfox746@gmail.com", p, msg)

            server.quit()

            print('Check your mail for OTP')

            while j < 3:
                o1 = int(input('Enter the OTP: '))
                if o1 == o:
                    p1 = input('Enter your new email: ')
                    ws[r1] = p1
                    print("Email changed successfully")
                    j = 3
                    i = 3
                else:
                    if j != 2:
                        print('Incorrect OTP')
                        j += 1
                    else:
                        print('Maximum number of attempts reached. ')
                        exit()

        else:
            print('Incorrect email')
            i += 1

    wb.save(filename='personal_data.xlsx')
    wb.close()

else:
    print("No email for this personnel")