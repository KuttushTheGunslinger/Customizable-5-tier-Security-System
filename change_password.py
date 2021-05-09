import smtplib
import random
from openpyxl import load_workbook

i, j = 0, 0

wb = load_workbook('personal_data.xlsx')
ws = wb.active

n = input('Enter your name: ')
l = []
for cell in ws['A']:
    l.append(cell.value)

if n in l:
    r = l.index(n) + 1
else:
    print("User doesn't exist")
    exit()

r1 = 'B' + str(r)

if ws[r1].value is not None:
    while i < 3:
        p = input('Enter your current password: ')
        if p == ws[r1].value:
                p1 = input('Enter your new password: ')
                ws[r1] = p1
                print("Password changed successfully")
                i = 3
        else:
            print('Incorrect password')
            i += 1

    wb.save(filename='personal_data.xlsx')
    wb.close()

else:
    print("No password for this personnel")