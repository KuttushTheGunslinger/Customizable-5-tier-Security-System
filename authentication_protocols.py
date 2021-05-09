import cv2
import numpy as np
import os
import random
import smtplib
import tensorflow as tf
from datetime import *
from openpyxl import load_workbook
from pytz import timezone
# import hashlib
# from pyfingerprint.pyfingerprint import PyFingerprint
# from pyfingerprint.pyfingerprint import FINGERPRINT_CHARBUFFER1

wb = load_workbook('personal_data.xlsx')
ws = wb.active

wb1 = load_workbook('system_info.xlsx')
ws1 = wb1.active

haar = 'haarcascade_frontalface_default.xml'
data = ['level_1', 'level_2']
(width, height) = (130, 100)


def password(n):
    k = 0
    a = 1
    for cell in ws['A']:
        if cell.value is not None:
            if n in cell.value:
                r = cell.row

    r1 = 'B' + str(r)

    while k < 3:
        p = input('Enter your password: ')

        if p == ws[r1].value:
            k = 3
        else:
            if k != 2:
                print('Incorrect password')
                k += 1
            else:
                print('Maximum number of attempts reached. ')
                exit()


def otp(n):
    for cell in ws['A']:
        if cell.value is not None:
            if n in cell.value:
                r = cell.row

    r2 = 'C' + str(r)

    j = 0
    o = random.randint(1000, 9999)

    server = smtplib.SMTP('smtp.gmail.com', port=587)
    server.starttls()
    server.login("gajeelredfox746@gmail.com", "passWORD123")

    msg = str(o)

    server.sendmail("gajeelredfox746@gmail.com", ws[r2].value, msg)

    server.quit()

    print('Check your mail for OTP')

    while j < 3:
        o1 = int(input('Enter the OTP: '))
        if o1 == o:
            j = 3
        else:
            if j != 2:
                print('Incorrect OTP')
                j += 1
            else:
                print('Maximum number of attempts reached. ')
                exit()

# def fingerprint_testing(Name):
#     l = []
#     try:
#         f = PyFingerprint('/dev/ttyUSB0', 57600, 0xFFFFFFFF, 0x00000000)
#
#         if ( f.verifyPassword() == False ):
#             raise ValueError('The given fingerprint sensor password is wrong!')
#
#     except Exception as e:
#         print('The fingerprint sensor could not be initialized!')
#         print('Exception message: ' + str(e))
#         exit(1)
#
#     ## Gets some sensor information
#     print('Currently used templates: ' + str(f.getTemplateCount()) +'/'+ str(f.getStorageCapacity()))
#
#     ## Tries to search the finger and calculate hash
#     try:
#         print('Waiting for finger...')
#
#         ## Wait that finger is read
#         while ( f.readImage() == False ):
#             pass
#
#         ## Converts read image to characteristics and stores it in charbuffer 1
#         f.convertImage(FINGERPRINT_CHARBUFFER1)
#
#         ## Searchs template
#         result = f.searchTemplate()
#
#         positionNumber = result[0]
#         accuracyScore = result[1]
#
#         if ( positionNumber == -1 ):
#             print('No match found!')
#             exit(0)
#         else:
#             wb3 = load_workbook('fingerprints.xlsx')
#             ws3 = wb3.active
#
#             for cell in ws3['A']:
#                 l.append(cell.value)
#
#             if Name in l:
#                 r = l.index(Name) + 1
#                 print(r)
#
#             if positionNumber == r - 2:
#                 print("Fingerprint matched")
#                 print('Found template at position #' + str(positionNumber))
#                 print('The accuracy score is: ' + str(accuracyScore))
#
# #
#         wb3.save(filename='fingerprints.xlsx')
#         wb3.close()
#
#
#     except Exception as e:
#         print('Operation failed!')
#         print('Exception message: ' + str(e))
#         exit(1)

    
def emotion():
    cap = cv2.VideoCapture(0)
    ret, image = cap.read()
    c = 1
    while c:
        gray1 = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        faces1 = face_cascade.detectMultiScale(gray1, 1.3, 5)
        for (x1, y1, w1, h1) in faces1:
            cv2.rectangle(image, (x1, y1), (x1 + w1, y1 + h1), (255, 0, 0), 2)
            face1 = gray1[y1:y1 + h1, x1:x1 + w1]
            face_resize1 = cv2.resize(face1, (width, height))
            cv2.imshow("face_resize1", face_resize1)
            cropped_img = np.expand_dims(np.expand_dims(cv2.resize(face_resize1, (48, 48)), -1), 0)

            my_model = tf.keras.models.load_model('model.h5')
            emotion_dict = {0: "Angry", 1: "Disgusted", 2: "Fearful", 3: "Happy", 4: "Neutral", 5: "Sad", 6: "Surprised"}
            prediction = my_model.predict(cropped_img)
            maxindex = int(np.argmax(prediction))
            if emotion_dict[maxindex] in ['Happy', 'Neutral']:
                print("Personnel is ", emotion_dict[maxindex])
                c = 0
                cap.release()
            else:
                cv2.putText(im, emotion_dict[maxindex], (x - 10, y - 10), cv2.FONT_HERSHEY_PLAIN, 1, (0, 255, 0))
                print('Personnel seems ' + emotion_dict[maxindex] + '. Not safe to give access')
                

print('Recognizing face please be in sufficient lights...')
(images, lables, names, id) = ([], [], {}, 0)

for d in data:
    for (subdirs, dirs, files) in os.walk(d):
        for subdir in dirs:
            names[id] = subdir
            subjectpath = os.path.join(d, subdir)
            for filename in os.listdir(subjectpath):
                path = subjectpath + '/' + filename
                lable = id
                images.append(cv2.imread(path, 0))
                lables.append(int(lable))
            id += 1
(images, lables) = [np.array(lis) for lis in [images, lables]]

model = cv2.face.LBPHFaceRecognizer_create()
model.train(images, lables)

face_cascade = cv2.CascadeClassifier(haar)

while True:
    webcam = cv2.VideoCapture(0)
    (_, im) = webcam.read()
    gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break
    faces = face_cascade.detectMultiScale(gray, 1.3, 5)
    for (x, y, w, h) in faces:
        cv2.rectangle(im, (x, y), (x + w, y + h), (255, 0, 0), 2)
        face = gray[y:y + h, x:x + w]
        face_resize = cv2.resize(face, (width, height))
        cv2.imshow("frame1", im)
        pred = model.predict(face_resize)
        print(pred)
        cv2.rectangle(im, (x, y), (x + w, y + h), (0, 255, 0), 3)
        webcam.release()
        if pred[1] < 100:
            print('Face recognized.')
            print('Hello ', names[pred[0]])
            if names[pred[0]] in os.listdir('level_1'):
                if ws1['B2'].value == 1:
                    emotion()
                # if ws1['B3'].value == 1:
                #     fingerprint_testing(names[pred[0]])
                if ws1['B4'].value == 1:
                    password(names[pred[0]])
                if ws1['B5'].value == 1:
                    otp(names[pred[0]])

                now = datetime.now(timezone('UTC'))
                now_asia = now.astimezone(timezone('Asia/Kolkata'))
                wb2 = load_workbook('entry_log.xlsx')
                ws2 = wb2.active

                r3 = ws2.max_row + 1
                ws2.cell(column=1, row=r3, value=names[pred[0]])
                ws2.cell(column=2, row=r3, value=now_asia.strftime("%d/%m/%Y"))
                ws2.cell(column=3, row=r3, value=now_asia.strftime("%H:%M:%S"))

                wb2.save(filename='entry_log.xlsx')
                wb2.close()

            else:
                if ws1['B8'].value == 1:
                    emotion()
                # if ws1['B9'].value == 1:
                #     fingerprint_testing(names[pred[0]])
                if ws1['B10'].value == 1:
                    password(names[pred[0]])
                if ws1['B11'].value == 1:
                    otp(names[pred[0]])

                now = datetime.now(timezone('UTC'))
                now_asia = now.astimezone(timezone('Asia/Kolkata'))
                wb2 = load_workbook('entry_log.xlsx')
                ws2 = wb2.active

                r3 = ws2.max_row + 1
                ws2.cell(column=1, row=r3, value=names[pred[0]])
                ws2.cell(column=2, row=r3, value=now_asia.strftime("%d/%m/%Y"))
                ws2.cell(column=3, row=r3, value=now_asia.strftime("%H:%M:%S"))

                wb2.save(filename='entry_log.xlsx')
                wb2.close()


