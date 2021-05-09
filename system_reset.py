import shutil
from openpyxl import load_workbook
import os

wb = load_workbook('system_info.xlsx')
ws = wb.active

wb1 = load_workbook('personal_data.xlsx')
ws1 = wb1.active

wb2 = load_workbook('fingerprints.xlsx')
ws2 = wb2.active

while 1:
    n = input("Do you want to delete this security system? (y/n): ")
    if n.lower() == 'y':
        for i in [1, 2, 3, 4, 5, 8, 9, 10, 11]:
            ws.cell(column=2, row=i, value=0)
            wb.save(filename='system_info.xlsx')
            wb.close()

        row = ws1.max_row
        while row > 1:
            ws1.delete_rows(row, 1)
            row = row - 1
            wb1.save(filename='personal_data.xlsx')
            wb1.close()

        row1 = ws2.max_row
        while row1 > 1:
            ws2.delete_rows(row1, 1)
            row1 = row1 - 1
            wb2.save(filename='fingerprints.xlsx')
            wb2.close()

        for f1 in os.listdir('level_1'):
            location1 = "D:/security/level_1/"
            path1 = os.path.join(location1, f1)
            shutil.rmtree(path1)

        for f2 in os.listdir('level_2'):
            location2 = "D:/security/level_2/"
            path2 = os.path.join(location2, f2)
            shutil.rmtree(path2)

        exit()

    elif n.lower() == 'n':
        exit()
    else:
        print("Invalid input. Please enter only y or n.")
