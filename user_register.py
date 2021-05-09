import cv2, os
from openpyxl import load_workbook
import time
# from pyfingerprint.pyfingerprint import PyFingerprint
# from pyfingerprint.pyfingerprint import FINGERPRINT_CHARBUFFER1
# from pyfingerprint.pyfingerprint import FINGERPRINT_CHARBUFFER2

wb = load_workbook('personal_data.xlsx')
ws = wb.active
row = ws.max_row + 1


def create_dataset(n, d):
    haar = 'haarcascade_frontalface_default.xml'
    path = os.path.join(d, n)
    if not os.path.isdir(path):
        os.mkdir(path)

    (width, height) = (130, 100)

    face_cascade = cv2.CascadeClassifier(haar)
    webcam = cv2.VideoCapture(0)

    count = 1
    while count < 100:
        (_, im) = webcam.read()
        gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.3, 4)
        for (x, y, w, h) in faces:
            cv2.rectangle(im, (x, y), (x + w, y + h), (255, 0, 0), 2)
            face = gray[y:y + h, x:x + w]
            face_resize = cv2.resize(face, (width, height))
            cv2.imwrite('% s/% s.png' % (path, count), face_resize)
        count += 1

        cv2.imshow('OpenCV', im)
        key = cv2.waitKey(10)
        if key == 27:
            break


def set_password():
    p = input("Enter your password: ")
    ws.cell(column=2, row=row, value=p)


def set_email():
    e = input("Enter your email: ")
    no = input("Enter your mobile number: ")
    ws.cell(column=3, row=row, value=e)
    ws.cell(column=4, row=row, value=no)


# def fingerprint(Name):
#     try:
#         f = PyFingerprint('/dev/ttyUSB0', 57600, 0xFFFFFFFF, 0x00000000)
# #
#         if (f.verifyPassword() == False):
#             raise ValueError('The given fingerprint sensor password is wrong!')
# #
#     except Exception as e:
#         print('The fingerprint sensor could not be initialized!')
#         print('Exception message: ' + str(e))
#         exit(1)
# #
#     ## Gets some sensor information
#     print('Currently used templates: ' + str(f.getTemplateCount()) + '/' + str(f.getStorageCapacity()))
# #
#     ## Tries to enroll new finger
#     try:
#         print('Waiting for finger...')
# #
#         ## Wait that finger is read
#         while (f.readImage() == False):
#             pass
# #
#         ## Converts read image to characteristics and stores it in charbuffer 1
#         f.convertImage(FINGERPRINT_CHARBUFFER1)
# #
#         ## Checks if finger is already enrolled
#         result = f.searchTemplate()
#         positionNumber = result[0]
# #
#         if (positionNumber >= 0):
#             print('Template already exists at position #' + str(positionNumber))
#             exit(0)
# #
#         print('Remove finger...')
#         time.sleep(2)
# #
#         print('Waiting for same finger again...')
# #
#         ## Wait that finger is read again
#         while (f.readImage() == False):
#             pass
# #
#         ## Converts read image to characteristics and stores it in charbuffer 2
#         f.convertImage(FINGERPRINT_CHARBUFFER2)
#
#         ## Compares the charbuffers
#         if (f.compareCharacteristics() == 0):
#             raise Exception('Fingers do not match')
# #
#         ## Creates a template
#         f.createTemplate()
# #
#         ## Saves template at new position number
#         positionNumber = f.storeTemplate()
#         print('Finger enrolled successfully!')
#         print('New template position #' + str(positionNumber))
# #
#     except Exception as e:
#         print('Operation failed!')
#         print('Exception message: ' + str(e))
#         exit(1)
# #
#     wb2 = load_workbook('fingerprints.xlsx')
#     ws2 = wb2.active
# #
#     row2 = ws2.max_row + 1
#     ws2.cell(column=1, row=row2, value=Name)
# #
#     wb2.save(filename='fingerprints.xlsx')
#     wb2.close()

wb1 = load_workbook('system_info.xlsx')
ws1 = wb1.active

name = input("Enter your name: ")
ws.cell(column=1, row=row, value=name)

if ws1['B1'] == 1:
    create_dataset(name, 'level_1')
    # if ws1['B3'].value == 1:
    #     fingerprint(name)
    if ws1['B4'].value == 1:
        set_password()
    if ws1['B5'].value == 1:
        set_email()

else:
    a = 1
    while a:
        m = input('Enter your clearance level (1/2): ')
        if m == '1':
            cl = 'level_1'
            a = 0
        elif m == '2':
            cl = 'level_2'
            a = 0
        else:
            print('Invalid')

    if cl == 'level_1':
        create_dataset(name, cl)
        # if ws1['B3'].value == 1:
        #     fingerprint(name)
        if ws1['B4'].value == 1:
            set_password()
        if ws1['B5'].value == 1:
            set_email()

    if cl == 'level_2':
        create_dataset(name, cl)
        # if ws1['B9'].value == 1:
        #     fingerprint(name)
        if ws1['B10'].value == 1:
            set_password()
        if ws1['B11'].value == 1:
            set_email()

wb.save(filename='personal_data.xlsx')
wb.close()

wb1.save(filename='system_info.xlsx')
wb1.close()

