from openpyxl import load_workbook

wb = load_workbook('system_info.xlsx')
ws = wb.active


def prompt():
    print("""
    1. Emotion detection
    2. Fingerprint verification
    3. Password verification
    4. One Time Password (OTP) verification

    Press 1 to enable, 0 to disable, e to exit.
    """)

while 1:
    print("How many levels of personnel do you want? (1/2): ")
    response = input("Enter your choice (press e to exit): ")
    if response == '1' or response == '2':
        break
    elif response.lower() == 'e':
        exit()
    else:
        print("Invalid input. Please enter only 1, 2, or e.\n")

a, b = 1, 1
if response == '1':
    ws.cell(column=2, row=1, value=1)
    print("What security protocols do you want for level 1?")
    prompt()
    while a:
        response1 = input("Enter your choice: ")
        if len(response1) != 4:
            print('Invalid input. Enter only 4 digits of 0s and 1s.\n')
            continue
        for i in response1:
            if i not in ['1', '0']:
                print("Invalid input. Please enter only 0s and 1s.\n")
                break
        else:
            a = 0

    if response1[0] == '1':
        ws.cell(column=2, row=2, value=1)
    if response1[1] == '1':
        ws.cell(column=2, row=3, value=1)
    if response1[2] == '1':
        ws.cell(column=2, row=4, value=1)
    if response1[3] == '1':
        ws.cell(column=2, row=5, value=1)


else:
    ws.cell(column=2, row=1, value=2)
    print("What security protocols do you want for level 1?")
    prompt()
    while a:
        response1 = input("Enter your choice: ")
        if len(response1) != 4:
            print('Invalid input. Enter only 4 digits of 0s and 1s.\n')
            continue
        for i in response1:
            if i not in ['1', '0']:
                print("Invalid input. Please enter only 0s and 1s.\n")
                break
        else:
            a = 0

    if response1[0] == '1':
        ws.cell(column=2, row=2, value=1)
    if response1[1] == '1':
        ws.cell(column=2, row=3, value=1)
    if response1[2] == '1':
        ws.cell(column=2, row=4, value=1)
    if response1[3] == '1':
        ws.cell(column=2, row=5, value=1)

    print("What security protocols do you want for level 2?")
    prompt()
    while b:
        response2 = input("Enter your choice: ")
        if len(response2) != 4:
            print('Invalid input. Enter only 4 digits of 0s and 1s.\n')
            continue
        for i in response2:
            if i not in ['1', '0']:
                print("Invalid input. Please enter only 0s and 1s.\n")
                break
        else:
            b = 0

    if response2[0] == '1':
        ws.cell(column=2, row=8, value=1)
    if response2[1] == '1':
        ws.cell(column=2, row=9, value=1)
    if response2[2] == '1':
        ws.cell(column=2, row=10, value=1)
    if response2[3] == '1':
        ws.cell(column=2, row=11, value=1)

wb.save(filename='system_info.xlsx')
wb.close()


